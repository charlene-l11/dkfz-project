import unittest
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from median_pipeline.excel_outputs import write_split_workbooks
from median_pipeline.preparation import discover, prepare_data, split_manifest, write_prepared


class SplittingTests(unittest.TestCase):
    def test_discovery_uses_configured_hypergui_and_labelling_patterns(self):
        import tempfile
        from pathlib import Path

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            experiment_folders = {}
            for label, subject in (("1_stom_pig", "P001"), ("2_liv_pig", "P002")):
                organ_root = root / label
                sample_dir = organ_root / f"{subject}_Experiment1" / "2026_01_01_00_00_00"
                hypergui_dir = sample_dir / "_hypergui_1"
                hypergui_dir.mkdir(parents=True)
                (sample_dir / "_labelling_001.txt").write_text("configured", encoding="utf-8")
                spectrum = hypergui_dir / "spectrum.csv"
                spectrum.write_text(
                    "\n".join(f"{500 + index * 5},{index / 100}" for index in range(100)),
                    encoding="utf-8",
                )

                # A configured HyperGUI without the configured labelling file is ignored
                # before its invalid spectrum is read.
                ignored_sample = organ_root / f"{subject}9_Experiment1" / "2026_01_02_00_00_00"
                ignored_hypergui = ignored_sample / "_hypergui_1"
                ignored_hypergui.mkdir(parents=True)
                (ignored_hypergui / "spectrum.csv").write_text("not,a,spectrum", encoding="utf-8")

                # A folder whose name does not match the configured HyperGUI pattern is ignored.
                wrong_sample = organ_root / f"{subject}8_Experiment1" / "2026_01_03_00_00_00"
                wrong_hypergui = wrong_sample / "_different_gui"
                wrong_hypergui.mkdir(parents=True)
                (wrong_sample / "_labelling_001.txt").write_text("configured", encoding="utf-8")
                (wrong_hypergui / "spectrum.csv").write_text("not,a,spectrum", encoding="utf-8")
                experiment_folders[label] = str(organ_root)

            cfg = {
                "experiment_folders": experiment_folders,
                "labelling_file": {
                    "1_stom_pig": "_labelling_*.txt",
                    "2_liv_pig": "_labelling_*.txt",
                },
                "hyperguis": {
                    "1_stom_pig": "_hypergui_*",
                    "2_liv_pig": "_hypergui_*",
                },
                "data": {
                    "patterns": ["spectrum.csv"],
                    "expected_channels": 100,
                    "annotation_name": "csv_spectrum",
                },
                "wavelength": {"min": 500, "max": 995, "selective": None},
            }
            frame, mapping, _ = discover(cfg)
            self.assertEqual(mapping, {"1_stom_pig": 0, "2_liv_pig": 1})
            self.assertEqual(len(frame), 2)
            self.assertTrue(frame["hypergui_dir"].str.endswith("_hypergui_1").all())
            self.assertTrue(frame["labelling_file"].str.endswith("_labelling_001.txt").all())

    def test_l1pixel_discovery_requires_the_l1pixel_folder(self):
        import tempfile
        from pathlib import Path

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            experiment_folders = {}
            filename = "spectrum_fromCSV1_masked_data_L1pixel_0_derivative.csv"
            spectrum_text = "\n".join(f"{500 + index * 5},{index / 100}" for index in range(100))
            for label, subject in (("stomach", "P001"), ("liver", "P002")):
                organ_root = root / label
                sample_dir = organ_root / f"{subject}_Experiment1" / "2026_01_01_00_00_00"
                hypergui_dir = sample_dir / "_hypergui_1"
                l1pixel_dir = hypergui_dir / "_L1pixel"
                l1pixel_dir.mkdir(parents=True)
                (sample_dir / "_labelling_001.txt").write_text("configured", encoding="utf-8")
                (l1pixel_dir / filename).write_text(spectrum_text, encoding="utf-8")
                (hypergui_dir / filename).write_text(spectrum_text, encoding="utf-8")
                experiment_folders[label] = str(organ_root)

            cfg = {
                "experiment_folders": experiment_folders,
                "labelling_file": {"stomach": "_labelling_001.txt", "liver": "_labelling_001.txt"},
                "hyperguis": {"stomach": "_hypergui_1", "liver": "_hypergui_1"},
                "data": {
                    "patterns": [f"_L1pixel/{filename}"],
                    "expected_channels": 100,
                    "annotation_name": "csv_spectrum",
                },
                "wavelength": {"min": 500, "max": 995, "selective": None},
            }

            frame, _, _ = discover(cfg)

            self.assertEqual(len(frame), 2)
            self.assertTrue(frame["file_path"].map(lambda value: Path(value).parent.name).eq("_L1pixel").all())

    def test_subjects_do_not_cross_splits_and_classes_are_covered(self):
        rows = []
        for subject in [f"P{i:03d}" for i in range(1, 9)]:
            for label in ("kidney", "liver"):
                rows.append({"subject_name": subject, "label": label})
        frame = pd.DataFrame(rows)
        cfg = {
            "splitting": {
                "train_ratio": 0.5,
                "val_ratio": 0.25,
                "test_ratio": 0.25,
                "seed": 42,
                "search_attempts": 1000,
                "require_all_classes_in": ["train", "val", "test"],
            }
        }
        result, summary = split_manifest(frame, cfg)
        assignments = result.groupby("subject_name")["split"].nunique()
        self.assertTrue((assignments == 1).all())
        for name in ("train", "val", "test"):
            self.assertEqual(set(result.loc[result["split"].eq(name), "label"]), {"kidney", "liver"})
            self.assertGreater(summary["splits"][name]["rows"], 0)

    def test_impossible_class_coverage_fails(self):
        frame = pd.DataFrame([
            {"subject_name": "P001", "label": "rare"},
            {"subject_name": "P002", "label": "rare"},
            {"subject_name": "P003", "label": "common"},
            {"subject_name": "P004", "label": "common"},
            {"subject_name": "P005", "label": "common"},
        ])
        cfg = {"splitting": {"train_ratio": 0.6, "val_ratio": 0.2, "test_ratio": 0.2, "seed": 1,
                              "search_attempts": 100, "require_all_classes_in": ["train", "val", "test"]}}
        with self.assertRaises(ValueError):
            split_manifest(frame, cfg)

    def test_external_testing_normalizes_primary_to_train_and_validation(self):
        rows = [
            {"subject_name": subject, "label": label}
            for subject in [f"P{i:03d}" for i in range(1, 7)]
            for label in ("kidney", "liver")
        ]
        cfg = {
            "external_testing": {"enabled": True},
            "splitting": {
                "train_ratio": 0.7,
                "val_ratio": 0.15,
                "test_ratio": 0.15,
                "seed": 42,
                "search_attempts": 1000,
                "require_all_classes_in": ["train", "val", "test"],
            },
        }

        result, summary = split_manifest(pd.DataFrame(rows), cfg)

        self.assertEqual(set(result["split"]), {"train", "val"})
        self.assertNotIn("test", summary["splits"])
        self.assertAlmostEqual(summary["effective_ratios"]["train"], 0.7 / 0.85)
        self.assertAlmostEqual(summary["effective_ratios"]["val"], 0.15 / 0.85)
        self.assertEqual(summary["effective_ratios"]["test"], 0)

    @patch("median_pipeline.preparation.discover")
    def test_prepare_data_assigns_second_dataset_only_to_test(self, discover_mock):
        primary = pd.DataFrame([
            {"subject_name": subject, "label": label}
            for subject in ("P001", "P002", "P003", "P004")
            for label in ("kidney", "liver")
        ])
        external = pd.DataFrame([
            {"subject_name": "M001", "label": "rat_muscle"},
            {"subject_name": "M001", "label": "rat_skin"},
        ])
        wavelength_info = {
            "selected_indices": [0, 1],
            "available_wavelengths": [500.0, 505.0],
        }
        discover_mock.side_effect = [
            (primary, {"kidney": 0, "liver": 1}, wavelength_info),
            (external, {"rat_muscle": 0, "rat_skin": 1}, wavelength_info),
        ]
        cfg = {
            "external_testing": {
                "enabled": True,
                "experiment_folders": {"rat_muscle": "/external-rat/muscle", "rat_skin": "/external-rat/skin"},
                "labelling_file": {"rat_muscle": "labels.txt", "rat_skin": "labels.txt"},
                "hyperguis": {"rat_muscle": "_hypergui_1", "rat_skin": "_hypergui_1"},
            },
            "splitting": {
                "train_ratio": 0.7,
                "val_ratio": 0.15,
                "test_ratio": 0.15,
                "seed": 42,
                "search_attempts": 100,
                "require_all_classes_in": ["train", "val", "test"],
            },
        }

        combined, mapping, summary, _ = prepare_data(cfg)

        self.assertEqual(mapping, {"kidney": 0, "liver": 1})
        self.assertEqual(set(combined.loc[combined["dataset_role"].eq("training_validation"), "split"]), {"train", "val"})
        self.assertTrue(combined.loc[combined["dataset_role"].eq("external_testing"), "split"].eq("test").all())
        self.assertEqual(summary["splits"]["test"]["source"], "external_testing")
        self.assertEqual(
            summary["external_testing_details"]["label_mapping"],
            {"rat_muscle": 0, "rat_skin": 1},
        )
        self.assertEqual(discover_mock.call_count, 2)
        testing_cfg = discover_mock.call_args_list[1].args[0]
        self.assertEqual(testing_cfg["experiment_folders"], cfg["external_testing"]["experiment_folders"])
        self.assertFalse(testing_cfg["external_testing"]["enabled"])

    def test_external_labels_are_written_with_an_independent_mapping(self):
        import json
        import tempfile
        from pathlib import Path

        rows = []
        for split, label, label_index, role in (
            ("train", "pig_liver", 0, "training_validation"),
            ("val", "pig_liver", 0, "training_validation"),
            ("test", "rat_fat", 0, "external_testing"),
        ):
            rows.append({
                "file_path": f"/{split}.csv",
                "label": label,
                "label_index": label_index,
                "subject_name": f"{split}_subject",
                "timestamp": "2026_01_01_00_00_00",
                "image_name": split,
                "annotation_name": "csv_spectrum",
                "dataset_role": role,
                "split": split,
            })
        summary = {
            "external_testing_details": {
                "labels": ["rat_fat"],
                "label_mapping": {"rat_fat": 0},
            }
        }

        with tempfile.TemporaryDirectory() as directory:
            data_dir = Path(directory)
            write_prepared(
                pd.DataFrame(rows),
                {"pig_liver": 0},
                summary,
                {"selected_indices": [0]},
                data_dir,
            )
            labels = json.loads((data_dir / "labels.json").read_text(encoding="utf-8"))
            test_manifest = pd.read_csv(data_dir / "test.csv")

            self.assertEqual(labels["mapping"], {"pig_liver": 0})
            self.assertEqual(labels["external_test_mapping"], {"rat_fat": 0})
            self.assertEqual(test_manifest.loc[0, "label"], "rat_fat")
            self.assertEqual(test_manifest.loc[0, "label_index"], 0)

    def test_per_organ_split_workbook_has_three_detail_sheets(self):
        import tempfile
        from pathlib import Path

        frame = pd.DataFrame([
            {"file_path": str(Path("P001") / "2026_01_01_00_00_00" / "_hypergui_1" / "spectrum.csv"),
             "label": "bladder", "subject_name": "P001", "individual_name": "P001_Experiment1",
             "timestamp": "2026_01_01_00_00_00", "image_name": "P001#2026_01_01_00_00_00", "split": split}
            for split in ("train", "val", "test")
        ])
        with tempfile.TemporaryDirectory() as directory:
            outputs = write_split_workbooks(frame, Path(directory))
            self.assertEqual([path.name for path in outputs], ["bladder_splits.xlsx"])
            workbook = load_workbook(outputs[0], data_only=True)
            self.assertEqual(workbook.sheetnames, ["Train Details", "Validation Details", "Test Details"])
            self.assertEqual(workbook["Train Details"]["E3"].value, 1)


if __name__ == "__main__":
    unittest.main()
