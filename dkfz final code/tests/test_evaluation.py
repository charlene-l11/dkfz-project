import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from median_pipeline.evaluation import write_evaluation, write_training_evaluation


class EvaluationTests(unittest.TestCase):
    def test_training_matrix_outputs_are_restricted(self):
        predictions = pd.DataFrame({
            "label_index": [0, 0, 1, 1],
            "prediction_index": [0, 1, 1, 1],
        })
        cfg = {"evaluation": {"formats": ["csv", "png", "pdf"], "dpi": 72}}

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            summary = write_training_evaluation(predictions, ["a", "b"], root, cfg)
            matrix_root = root / "training_matrices"
            expected = {
                Path("confusion_matrix.xlsx"),
                Path("csv/confusion_matrix_raw.csv"),
                Path("csv/confusion_matrix_normalized.csv"),
                Path("csv/per_class_metrics.csv"),
                Path("png/confusion_matrix_raw_light.png"),
                Path("png/confusion_matrix_normalized_light.png"),
            }
            actual = {
                path.relative_to(matrix_root)
                for path in matrix_root.rglob("*")
                if path.is_file()
            }

            self.assertEqual(actual, expected)
            self.assertFalse((root / "training_metrics.csv").exists())
            self.assertEqual(summary["n_training_rows"], 4)

    def test_rectangular_external_confusion_matrix(self):
        predicted_labels = [
            "stomach", "liver", "colon", "kidney", "spleen",
            "bladder", "skin", "muscle", "heart",
        ]
        true_labels = ["liver", "rat_fat", "kidney"]
        predictions = pd.DataFrame({
            "label_index": [0, 0, 1, 1, 2, 2],
            "label": ["liver", "liver", "rat_fat", "rat_fat", "kidney", "kidney"],
            "prediction_index": [1, 3, 7, 7, 3, 0],
            "prediction_label": ["liver", "kidney", "muscle", "muscle", "kidney", "stomach"],
        })
        cfg = {"evaluation": {"formats": ["csv"], "dpi": 72}}

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            summary = write_evaluation(
                predictions,
                predicted_labels,
                root,
                cfg,
                split_name="testing",
                true_labels=true_labels,
            )
            matrix = pd.read_csv(
                root / "testing_matrices/csv/confusion_matrix_raw.csv",
                index_col=0,
            )

            self.assertEqual(matrix.shape, (3, 9))
            self.assertEqual(matrix.index.tolist(), true_labels)
            self.assertEqual(matrix.columns.tolist(), predicted_labels)
            self.assertEqual(matrix.loc["rat_fat", "muscle"], 2)
            self.assertAlmostEqual(summary["accuracy"], 2 / 6)
            self.assertEqual(summary["n_true_classes"], 3)
            self.assertEqual(summary["n_predicted_classes"], 9)

    def test_all_matrix_and_metric_outputs_are_written(self):
        predictions = pd.DataFrame({
            "label_index": [0, 0, 1, 1],
            "prediction_index": [0, 1, 1, 1],
        })
        cfg = {"evaluation": {"formats": ["csv", "png", "pdf"], "dpi": 72}}
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            validation_summary = write_evaluation(
                predictions, ["a", "b"], root, cfg, split_name="validation"
            )
            summary = write_evaluation(predictions, ["a", "b"], root, cfg, split_name="testing")
            expected = [
                root / "confusion_matrix_normalized_light.pdf",
                root / "confusion_matrix_normalized_dark.pdf",
                root / "summary.json",
                root / "validation_metrics.csv",
                root / "testing_metrics.csv",
            ]
            for split_name in ("validation", "testing"):
                matrix_root = root / f"{split_name}_matrices"
                expected.extend([
                    matrix_root / "csv/confusion_matrix_raw.csv",
                    matrix_root / "csv/confusion_matrix_normalized.csv",
                    matrix_root / "csv/per_class_metrics.csv",
                    matrix_root / "png/confusion_matrix_raw_light.png",
                    matrix_root / "png/confusion_matrix_raw_dark.png",
                    matrix_root / "png/confusion_matrix_normalized_light.png",
                    matrix_root / "png/confusion_matrix_normalized_dark.png",
                    matrix_root / "pdf/confusion_matrix_raw_light.pdf",
                    matrix_root / "pdf/confusion_matrix_raw_dark.pdf",
                    matrix_root / "pdf/confusion_matrix_normalized_light.pdf",
                    matrix_root / "pdf/confusion_matrix_normalized_dark.pdf",
                    matrix_root / "confusion_matrix.xlsx",
                ])
            self.assertTrue(all(path.exists() for path in expected))
            self.assertEqual(summary["accuracy"], 0.75)
            self.assertEqual(validation_summary["accuracy"], 0.75)
            workbook = load_workbook(root / "testing_matrices/confusion_matrix.xlsx", data_only=True)
            self.assertEqual(workbook.sheetnames, ["Matrix Normalized", "Matrix Counts"])
            self.assertAlmostEqual(workbook["Matrix Normalized"]["B4"].value, 0.5)
            self.assertEqual(workbook["Matrix Normalized"]["B4"].number_format, "0.000")
            self.assertEqual(workbook["Matrix Counts"]["B4"].value, 1)
            normalized_csv = (
                root / "testing_matrices/csv/confusion_matrix_normalized.csv"
            ).read_text(encoding="utf-8")
            self.assertIn("a,0.500,0.500", normalized_csv)
            self.assertIn("b,0.000,1.000", normalized_csv)
            self.assertNotIn("%", normalized_csv)
            validation_metrics = pd.read_csv(root / "validation_metrics.csv")
            testing_metrics = pd.read_csv(root / "testing_metrics.csv")
            self.assertAlmostEqual(validation_metrics.loc[0, "accuracy"], 0.75)
            self.assertAlmostEqual(testing_metrics.loc[0, "accuracy"], 0.75)
            self.assertEqual(validation_metrics.loc[0, "n_validation_rows"], 4)
            self.assertEqual(testing_metrics.loc[0, "n_test_rows"], 4)
            unexpected = [
                root / "testing_matrices/png/confusion_matrix_raw.png",
                root / "testing_matrices/png/confusion_matrix_normalized.png",
                root / "testing_matrices/pdf/confusion_matrix_raw.pdf",
                root / "testing_matrices/pdf/confusion_matrix_normalized.pdf",
                root / "confusion_matrix_normalized.pdf",
                root / "validation_matrices/summary.json",
                root / "validation_summary.json",
                root / "validation_matrices/metrics.csv",
                root / "testing_matrices/metrics.csv",
            ]
            self.assertTrue(all(not path.exists() for path in unexpected))


if __name__ == "__main__":
    unittest.main()
