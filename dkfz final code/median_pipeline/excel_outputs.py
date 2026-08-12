from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _status(message: str) -> None:
    print(message, flush=True)


def _safe_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._") or "organ"


def write_wavelength_workbook(wavelength_info: dict[str, Any], output: Path) -> Path:
    """Write a readable Excel copy of the resolved wavelength metadata."""
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview.sheet_view.showGridLines = False
    overview.freeze_panes = "A4"
    overview["A1"] = "Wavelength Selection"
    overview["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    overview["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    overview.merge_cells("A1:B1")
    for column, header in enumerate(("Field", "Value"), start=1):
        cell = overview.cell(row=3, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
        cell.alignment = Alignment(horizontal="center")

    scalar_items = [
        (key, value)
        for key, value in wavelength_info.items()
        if key not in {"available_wavelengths", "selected_indices", "selected_wavelengths"}
    ]
    for row_number, (key, value) in enumerate(scalar_items, start=4):
        overview.cell(row=row_number, column=1, value=key)
        if isinstance(value, (dict, list)):
            value = json.dumps(value, ensure_ascii=False)
        overview.cell(row=row_number, column=2, value=value)
    overview.auto_filter.ref = f"A3:B{max(3, 3 + len(scalar_items))}"
    overview.column_dimensions["A"].width = 28
    overview.column_dimensions["B"].width = 60

    wavelengths = workbook.create_sheet("Wavelengths")
    wavelengths.sheet_view.showGridLines = False
    wavelengths.freeze_panes = "A2"
    for column, header in enumerate(("Input index", "Wavelength (nm)", "Selected"), start=1):
        cell = wavelengths.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
        cell.alignment = Alignment(horizontal="center")
    selected_indices = {int(value) for value in wavelength_info.get("selected_indices", [])}
    available = wavelength_info.get("available_wavelengths", [])
    for row_number, value in enumerate(available, start=2):
        index = row_number - 2
        for column, cell_value in enumerate(
            (index, float(value), "Yes" if index in selected_indices else "No"), start=1
        ):
            cell = wavelengths.cell(row=row_number, column=column, value=cell_value)
            cell.alignment = Alignment(horizontal="center")
    wavelengths.auto_filter.ref = f"A1:C{max(1, len(available) + 1)}"
    wavelengths.column_dimensions["A"].width = 16
    wavelengths.column_dimensions["B"].width = 20
    wavelengths.column_dimensions["C"].width = 14

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output


def _detail_value(row: pd.Series, name: str) -> str:
    if name in row and pd.notna(row[name]):
        return str(row[name])
    path = Path(str(row["file_path"]))
    if name == "individual_name":
        return str(row["subject_name"])
    if name == "sample_dir":
        return str(path.parent.parent)
    if name == "hypergui_dir":
        return str(path.parent)
    if name == "labelling_file":
        return ""
    raise KeyError(name)


def _style_detail_sheet(sheet, title: str, subset: pd.DataFrame) -> None:
    sheet.title = title
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A7"
    sheet["A1"] = title
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet.merge_cells("A1:G1")
    individuals = subset.apply(lambda row: _detail_value(row, "individual_name"), axis=1) if len(subset) else []
    sheet["A3"] = "Unique individuals"
    sheet["B3"] = len(set(individuals))
    sheet["D3"] = "Unique recordings"
    sheet["E3"] = len(subset)
    sheet["A4"] = "Individual names are shown once, followed by the recordings assigned to this split."
    sheet.merge_cells("A4:G4")
    headers = [
        "individual / recording",
        "label_name",
        "image_name",
        "sample_dir",
        "hypergui_dir",
        "labelling_file",
        "spectrum_csv_path",
    ]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=6, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
        cell.alignment = Alignment(horizontal="center")

    rows = subset.copy()
    row_number = 7

    if not rows.empty:
        rows["_individual"] = rows.apply(
            lambda row: _detail_value(row, "individual_name"), axis=1
        )
        rows = rows.sort_values(["_individual", "timestamp", "file_path"])
        grouped_rows = rows.groupby("_individual", sort=True)
    else:
        grouped_rows = ()

    for individual, recordings in grouped_rows:
        sheet.cell(row=row_number, column=1, value=str(individual)).font = Font(bold=True)
        sheet.cell(row=row_number, column=2, value=f"{len(recordings):,} recordings").font = Font(bold=True)
        for column in range(1, 8):
            sheet.cell(row=row_number, column=column).fill = PatternFill("solid", fgColor="D9EAF7")
        row_number += 1
        for _, record in recordings.iterrows():
            values = [
                f"  {record['timestamp']}", str(record["label"]), str(record["image_name"]),
                _detail_value(record, "sample_dir"),
                _detail_value(record, "hypergui_dir"),
                _detail_value(record, "labelling_file"),
                str(record["file_path"]),
            ]
            for column, value in enumerate(values, start=1):
                sheet.cell(row=row_number, column=column, value=value)
            row_number += 1
    sheet.auto_filter.ref = f"A6:G{max(6, row_number - 1)}"
    for column, width in enumerate([34, 20, 55, 70, 70, 55, 90], start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width


def write_split_workbooks(frame: pd.DataFrame, splits_dir: Path) -> list[Path]:
    splits_dir.mkdir(parents=True, exist_ok=True)
    labels = sorted(frame["label"].astype(str).unique())
    _status(f"  Writing {len(labels)} per-organ split workbooks to {splits_dir}")
    outputs = []
    for number, label in enumerate(labels, start=1):
        workbook = Workbook()
        organ = frame[frame["label"].astype(str).eq(label)]
        for sheet_number, (split, title) in enumerate((("train", "Train Details"), ("val", "Validation Details"), ("test", "Test Details"))):
            sheet = workbook.active if sheet_number == 0 else workbook.create_sheet()
            _style_detail_sheet(sheet, title, organ[organ["split"].eq(split)])
        output = splits_dir / f"{_safe_filename(label)}_splits.xlsx"
        workbook.save(output)
        outputs.append(output)
        _status(f"      [{number}/{len(labels)}] Wrote {output.name}")
    return outputs


def _write_matrix_sheet(sheet, frame: pd.DataFrame, note: str, normalized: bool) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "B4"
    sheet["A1"] = note
    sheet["A1"].font = Font(italic=True)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(frame.columns) + 1)
    for column, value in enumerate(["true_label", *frame.columns.tolist()], start=1):
        cell = sheet.cell(row=3, column=column, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
        cell.alignment = Alignment(horizontal="center")
    for row_number, (label, values) in enumerate(frame.iterrows(), start=4):
        label_cell = sheet.cell(row=row_number, column=1, value=label)
        label_cell.font = Font(bold=True)
        label_cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for column, value in enumerate(values, start=2):
            cell = sheet.cell(row=row_number, column=column, value=float(value) if normalized else int(value))
            if normalized:
                cell.number_format = "0.000"
    last_row = len(frame) + 3
    last_column = len(frame.columns) + 1
    sheet.auto_filter.ref = f"A3:{get_column_letter(last_column)}{last_row}"
    sheet.column_dimensions["A"].width = 24
    for column in range(2, last_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 15


def write_matrix_workbook(raw_frame: pd.DataFrame, normalized_fraction: pd.DataFrame, output: Path) -> None:
    workbook = Workbook()
    normalized_sheet = workbook.active
    normalized_sheet.title = "Matrix Normalized"
    _write_matrix_sheet(normalized_sheet, normalized_fraction, "Normalized confusion matrix. Values are fractions of test recordings per true label.", True)
    counts_sheet = workbook.create_sheet("Matrix Counts")
    _write_matrix_sheet(counts_sheet, raw_frame, "Raw confusion matrix counts. Counts are test recordings/prediction rows, not individuals.", False)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
